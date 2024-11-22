# openpyxl读文件和写文件是分开的

from openpyxl import load_workbook  # 读文件


# data_only=True 会把表格中的计算公式直接展示
# 通过代码产生的表格需要先在表格文件中保存一次才能正常显示出公式的值
wb = load_workbook('test.xlsx',read_only=True,data_only=True)
print(wb)
# 查看有几个表格
print(wb.sheetnames)  # ['login', 'Sheet', 'index1']
print(wb['sheet1']['A1'].value)
print(wb['sheet1']['A2'].value)

print('='*100)

# 这种方法的取值是横纵坐标的最大值
res = wb['sheet2']
for row in res.rows:
    for j in row:
        print(j.value,end='|')
    print()