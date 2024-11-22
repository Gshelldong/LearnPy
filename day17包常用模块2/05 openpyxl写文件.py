# openpyxl模块 比较火的操作excel表的模块

"""
03版本之前 excel文件的后缀名 叫xls
03版本之后 excel文件的后缀名 叫xlsx

xlwd  写excel
xlrt  读excel

xlwd和xlrt既支持03版本之前的excel文件也支持03版本之后的excel文件
openpyxl 只支持03版本之后的  xlsx
"""

# 写表格

from openpyxl import Workbook

wb = Workbook() # 创建一个工作簿
wb_sheet1 = wb.create_sheet("sheet1",0) # 在第一个sheet的位置创件一个sheet1
wb_sheet2 = wb.create_sheet("sheet2",1)

# 写入内容
# 第一种写入方式
wb_sheet1['A1'] = '序号'
wb_sheet1['A2'] = 1
wb_sheet1['A3'] = 2

# 第二种写入方式
wb_sheet2.cell(row=1,column=1,value='名称')
wb_sheet2.cell(row=3,column=4,value='666')

# 再或者写入内容

wb_sheet2.append(['姓名','年龄','爱好'])  # 写表头
wb_sheet2.append(['tom','12','sports'])  #内容
wb_sheet2.append(['jerry','10',])  # 如果某个不想填写就写空值
wb_sheet2.append(['mario','','run'])

wb.save('test.xlsx')





